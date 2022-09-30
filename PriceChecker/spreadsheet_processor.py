"""This module just save new data to spreadsheet
and compare it to return that items if their cost is lower """
from datetime import datetime
import openpyxl

FIRST_ITEM_ROW = 3
SPREADSHEET = 'items.xlsx'
NUM_COL = 1
URL_COL = 2
DESCRIPTION_COL = 3


def save_prices_to_table(list_of_items, user_id):
    """Parse incoming dictionary with list of items {'URL to item':[item_description, item price in Rubles])"""
    wb = openpyxl.load_workbook(f'prices/{str(user_id)}/{SPREADSHEET}')
    sheet = wb.active

    max_col = sheet.max_column
    urls = list(list_of_items.keys())

    for url in range(len(urls)):
        sheet.cell(FIRST_ITEM_ROW + url, NUM_COL).value = url + 1
        sheet.cell(FIRST_ITEM_ROW + url, URL_COL).value = urls[url]

    for item in range(len(list_of_items)):
        sheet.cell(FIRST_ITEM_ROW + item, DESCRIPTION_COL).value = list_of_items.get(urls[item])[0]
        sheet.cell(FIRST_ITEM_ROW + item, max_col + 1).value = list_of_items.get(urls[item])[1]
        sheet.cell(FIRST_ITEM_ROW - 1, max_col + 1).value = datetime.date(datetime.today())
    wb.save(f'prices/{str(user_id)}/{SPREADSHEET}')


def compare_prices_with_the_past(user_id):
    """Check that prices in new colon is lower than prices in previous colon"""
    wb = openpyxl.load_workbook(f'prices/{str(user_id)}/{SPREADSHEET}')
    sheet = wb.active
    difference = []
    max_col = sheet.max_column
    max_row = sheet.max_row

    for item in range(0, max_row - FIRST_ITEM_ROW + 1):
        if sheet.cell(FIRST_ITEM_ROW + item, max_col-1).value is None:
            continue
        else:
            if sheet.cell(FIRST_ITEM_ROW + item, max_col).value < sheet.cell(FIRST_ITEM_ROW + item, max_col - 1).value:
                difference.append(f'Price for {sheet.cell(FIRST_ITEM_ROW + item, DESCRIPTION_COL).value} now is lower. \n'
                                 f'Previous Price {sheet.cell(FIRST_ITEM_ROW + item, max_col - 1).value} \n'
                                 f'New Price {sheet.cell(FIRST_ITEM_ROW + item, max_col).value} \n'
                                 f'Link to the Store: {sheet.cell(FIRST_ITEM_ROW + item, URL_COL).value})')
            elif sheet.cell(FIRST_ITEM_ROW + item, max_col).value > sheet.cell(FIRST_ITEM_ROW + item, max_col - 1).value:
                difference.append(f'Price for {sheet.cell(FIRST_ITEM_ROW + item, DESCRIPTION_COL).value} now is higher. \n'
                                 f'Previous Price {sheet.cell(FIRST_ITEM_ROW + item, max_col - 1).value} \n'
                                 f'New Price {sheet.cell(FIRST_ITEM_ROW + item, max_col).value} \n'
                                 f'Link to the Store: {sheet.cell(FIRST_ITEM_ROW + item, URL_COL).value})')
    return difference