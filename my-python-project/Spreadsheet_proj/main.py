import openpyxl

book_path = "Pricebook.xlsx"
book_obj = openpyxl.load_workbook(book_path)
sheet_obj = book_obj.active
max_row = sheet_obj.max_row
max_column = sheet_obj.max_column

# How much products each company have - new company can be added to Excel - it will count. Not only 3.
company_list = []
company_list_count = []
result = {}

for x in range(2, max_row+1):
    company_list.append(sheet_obj.cell(row=x, column=4).value)
company_list_unique = list(set(company_list))
for x in company_list_unique:
    counter = 0
    for y in company_list:
        if y == x:
            counter = counter+1
    company_list_count.append(counter)
for x in range (0, len(company_list_unique)):
    result[company_list_unique[x]] = company_list_count[x]
print(result)

# how much products with Inventory less 1000
invent_less_than = {}
for x in range(2, max_row+1):
    if sheet_obj.cell(row=x, column=2).value < 1000:
        invent_less_than[sheet_obj.cell(row=x, column=1).value] = sheet_obj.cell(row=x, column=2).value
print(invent_less_than)

# Inventory value of each company
company_list_value = []
for company in company_list_unique:
    counter = 0
    for x in range(2, max_row+1):
        if sheet_obj.cell(row=x, column=4).value == company:
            counter = counter+(sheet_obj.cell(row=x, column=2)).value*(sheet_obj.cell(row=x, column=3)).value
    company_list_value.append(counter)
for x in range(0, len(company_list_unique)):
    result[company_list_unique[x]] = company_list_value[x]
print(result)

# Write Value for each product
sheet_obj.cell(row=1, column=max_column+1).value = "Product value"
for x in range(2, max_row+1):
    sheet_obj.cell(row=x, column=max_column+1).value = (sheet_obj.cell(row=x, column=2)).value/(sheet_obj.cell(row=x, column=3)).value
book_obj.save("Pricebook2.xlsx")
